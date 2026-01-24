"""
Rutas para carga masiva desde archivos Excel.
"""
from io import BytesIO

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook, load_workbook

from api.middleware import requiere_autenticacion, requiere_rol
from modelos.base_datos import BaseDatos
from modelos.categoria_modelo import CategoriaModelo
from modelos.plan_modelo import PlanModelo
from modelos.producto_modelo import ProductoModelo
from modelos.salon_modelo import SalonModelo
from utilidades.logger import obtener_logger

carga_masiva_bp = Blueprint("carga_masiva", __name__)
logger = obtener_logger()

TIPOS_PLANTILLA = {
    "categorias": [
        "id",
        "nombre",
        "descripcion",
        "activo",
    ],
    "productos": [
        "id",
        "nombre",
        "descripcion",
        "detalles_adicionales",
        "variantes",
        "precio",
        "precio_minimo",
        "precio_maximo",
        "duracion_horas",
        "categoria",
        "id_categoria",
        "stock",
        "unidad_medida",
        "tipo_servicio",
        "activo",
    ],
    "salones": [
        "id_salon",
        "nombre",
        "capacidad",
        "ubicacion",
        "descripcion",
        "precio_base",
        "activo",
    ],
    "planes": [
        "id",
        "nombre",
        "descripcion",
        "precio_base",
        "capacidad_minima",
        "capacidad_maxima",
        "duracion_horas",
        "incluye",
        "activo",
    ],
}


def _normalizar_header(valor):
    return str(valor or "").strip().lower()


def _parse_bool(valor, default=True):
    if valor is None or str(valor).strip() == "":
        return default
    texto = str(valor).strip().lower()
    if texto in ("1", "true", "si", "sí", "activo", "yes"):
        return True
    if texto in ("0", "false", "no", "inactivo"):
        return False
    return default


def _parse_float(valor, default=0):
    if valor is None or str(valor).strip() == "":
        return default
    try:
        return float(str(valor).replace(",", "."))
    except ValueError:
        return default


def _parse_int(valor, default=None):
    if valor is None or str(valor).strip() == "":
        return default
    try:
        return int(float(valor))
    except ValueError:
        return default


@carga_masiva_bp.route("/template/<tipo>", methods=["GET"])
@requiere_autenticacion
@requiere_rol("administrador", "gerente_general")
def descargar_template(tipo):
    tipo = (tipo or "").strip().lower()
    columnas = TIPOS_PLANTILLA.get(tipo)
    if not columnas:
        return jsonify({"error": "Tipo de plantilla no válido"}), 400

    wb = Workbook()
    ws = wb.active
    ws.title = "Plantilla"
    ws.append(columnas)

    salida = BytesIO()
    wb.save(salida)
    salida.seek(0)

    nombre_archivo = f"plantilla_{tipo}.xlsx"
    return send_file(
        salida,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@carga_masiva_bp.route("/<tipo>", methods=["POST"])
@requiere_autenticacion
@requiere_rol("administrador", "gerente_general")
def cargar_excel(tipo):
    tipo = (tipo or "").strip().lower()
    columnas = TIPOS_PLANTILLA.get(tipo)
    if not columnas:
        return jsonify({"error": "Tipo de carga no válido"}), 400

    if "archivo" not in request.files:
        return jsonify({"error": "Archivo es requerido"}), 400

    archivo = request.files["archivo"]
    if not archivo or archivo.filename == "":
        return jsonify({"error": "Archivo inválido"}), 400

    try:
        wb = load_workbook(archivo, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error(f"Error al leer Excel: {e}")
        return jsonify({"error": "No se pudo leer el archivo Excel"}), 400

    filas = list(ws.iter_rows(values_only=True))
    if not filas:
        return jsonify({"error": "El archivo no contiene datos"}), 400

    headers = [_normalizar_header(h) for h in filas[0]]
    indices = {col: headers.index(col) for col in columnas if col in headers}
    if "nombre" not in indices:
        return jsonify({"error": "La plantilla no contiene la columna 'nombre'"}), 400

    resultados = {"creados": 0, "actualizados": 0, "omitidos": 0, "errores": []}

    categorias_modelo = CategoriaModelo()
    productos_modelo = ProductoModelo()
    salones_modelo = SalonModelo()
    planes_modelo = PlanModelo()
    base = BaseDatos()

    for idx, fila in enumerate(filas[1:], start=2):
        fila_dict = {col: (fila[indices[col]] if col in indices else None) for col in columnas}
        nombre = str(fila_dict.get("nombre") or "").strip()
        if not nombre:
            resultados["omitidos"] += 1
            continue
        try:
            if tipo == "categorias":
                categoria_id = _parse_int(fila_dict.get("id"))
                if not categoria_id:
                    consulta = "SELECT id FROM categorias WHERE LOWER(nombre) = LOWER(%s) LIMIT 1"
                    existente = base.obtener_uno(consulta, (nombre,))
                    categoria_id = existente["id"] if existente else None
                datos = {
                    "nombre": nombre,
                    "descripcion": fila_dict.get("descripcion"),
                    "activo": _parse_bool(fila_dict.get("activo"), True),
                }
                if categoria_id:
                    categorias_modelo.actualizar_categoria(categoria_id, datos)
                    resultados["actualizados"] += 1
                else:
                    categorias_modelo.crear_categoria(datos)
                    resultados["creados"] += 1

            elif tipo == "salones":
                salon_id = _parse_int(fila_dict.get("id_salon"))
                if not salon_id:
                    consulta = "SELECT id_salon FROM salones WHERE LOWER(nombre) = LOWER(%s) LIMIT 1"
                    existente = base.obtener_uno(consulta, (nombre,))
                    salon_id = existente["id_salon"] if existente else None
                datos = {
                    "nombre": nombre,
                    "capacidad": _parse_int(fila_dict.get("capacidad"), 0),
                    "ubicacion": fila_dict.get("ubicacion"),
                    "descripcion": fila_dict.get("descripcion"),
                    "precio_base": _parse_float(fila_dict.get("precio_base"), 0),
                    "activo": _parse_bool(fila_dict.get("activo"), True),
                }
                if salon_id:
                    salones_modelo.actualizar_salon(salon_id, datos)
                    resultados["actualizados"] += 1
                else:
                    salones_modelo.crear_salon(datos)
                    resultados["creados"] += 1

            elif tipo == "planes":
                plan_id = _parse_int(fila_dict.get("id"))
                if not plan_id:
                    consulta = "SELECT id FROM planes WHERE LOWER(nombre) = LOWER(%s) LIMIT 1"
                    existente = base.obtener_uno(consulta, (nombre,))
                    plan_id = existente["id"] if existente else None
                datos = {
                    "nombre": nombre,
                    "descripcion": fila_dict.get("descripcion"),
                    "precio_base": _parse_float(fila_dict.get("precio_base"), 0),
                    "capacidad_minima": _parse_int(fila_dict.get("capacidad_minima")),
                    "capacidad_maxima": _parse_int(fila_dict.get("capacidad_maxima")),
                    "duracion_horas": _parse_int(fila_dict.get("duracion_horas")),
                    "incluye": fila_dict.get("incluye"),
                    "activo": _parse_bool(fila_dict.get("activo"), True),
                }
                if plan_id:
                    planes_modelo.actualizar_plan(plan_id, datos)
                    resultados["actualizados"] += 1
                else:
                    planes_modelo.crear_plan(datos)
                    resultados["creados"] += 1

            elif tipo == "productos":
                producto_id = _parse_int(fila_dict.get("id"))
                id_categoria = _parse_int(fila_dict.get("id_categoria"))
                nombre_categoria = str(fila_dict.get("categoria") or "").strip()
                if not id_categoria and nombre_categoria:
                    consulta = "SELECT id FROM categorias WHERE LOWER(nombre) = LOWER(%s) LIMIT 1"
                    categoria = base.obtener_uno(consulta, (nombre_categoria,))
                    id_categoria = categoria["id"] if categoria else None
                if not id_categoria and nombre_categoria:
                    resultados["errores"].append(
                        {"fila": idx, "error": f"Categoría '{nombre_categoria}' no encontrada"}
                    )
                    continue
                if not producto_id:
                    if id_categoria:
                        consulta = "SELECT id FROM productos WHERE LOWER(nombre) = LOWER(%s) AND id_categoria = %s LIMIT 1"
                        existente = base.obtener_uno(consulta, (nombre, id_categoria))
                    else:
                        consulta = "SELECT id FROM productos WHERE LOWER(nombre) = LOWER(%s) LIMIT 1"
                        existente = base.obtener_uno(consulta, (nombre,))
                    producto_id = existente["id"] if existente else None

                datos = {
                    "nombre": nombre,
                    "descripcion": fila_dict.get("descripcion"),
                    "detalles_adicionales": fila_dict.get("detalles_adicionales"),
                    "variantes": fila_dict.get("variantes"),
                    "precio": _parse_float(fila_dict.get("precio"), 0),
                    "precio_minimo": _parse_float(fila_dict.get("precio_minimo"), None),
                    "precio_maximo": _parse_float(fila_dict.get("precio_maximo"), None),
                    "duracion_horas": _parse_int(fila_dict.get("duracion_horas")),
                    "id_categoria": id_categoria,
                    "stock": _parse_int(fila_dict.get("stock"), 0) or 0,
                    "unidad_medida": fila_dict.get("unidad_medida") or "unidad",
                    "tipo_servicio": fila_dict.get("tipo_servicio") or "servicio",
                    "activo": _parse_bool(fila_dict.get("activo"), True),
                }
                if producto_id:
                    productos_modelo.actualizar_producto(producto_id, datos)
                    resultados["actualizados"] += 1
                else:
                    productos_modelo.crear_producto(datos)
                    resultados["creados"] += 1

        except Exception as e:
            logger.error(f"Error en fila {idx}: {e}")
            resultados["errores"].append({"fila": idx, "error": str(e)})

    return jsonify(resultados), 200
